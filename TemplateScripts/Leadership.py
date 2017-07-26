import xlsxwriter, re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class Staff:
	name = ""
	gender = ""
	ethnicity = ""
	picture = ""

#open xlsx file
wb = load_workbook(filename='Consolidated Leadership Info.xlsx')
ws = wb['Sheet1']

#create new spreadsheet
workbook  = xlsxwriter.Workbook('leaders.xlsx')
worksheet = workbook.add_worksheet()
rowNum = 0
#initial title setup
worksheet.write(rowNum, 0, 'IPEDS code')
worksheet.write(rowNum, 1, 'Institution')
worksheet.write(rowNum, 2, '# WOC per institution')
worksheet.write(rowNum, 3, '# Engineering departments')
worksheet.write(rowNum, 4, 'President (or equivalent) Name')
worksheet.write(rowNum, 5, "President's gender")
worksheet.write(rowNum, 6, "President's ethnicity")
worksheet.write(rowNum, 7, "President picture link")
worksheet.write(rowNum, 8, "Provost (or equivalent) Name")
worksheet.write(rowNum, 9, 'Provost Gender')
worksheet.write(rowNum, 10, 'Provost Ethnicity')
worksheet.write(rowNum, 11, 'Provost picture link')
worksheet.write(rowNum, 12, 'Diversity officer name (or any equivalent)')
worksheet.write(rowNum, 13, 'Diversity officer (or equivalent) title')
worksheet.write(rowNum, 14, 'Diversity officer gender')
worksheet.write(rowNum, 15, 'Diversity officer ethnicity')
worksheet.write(rowNum, 16, 'Diversity officer picture link')
rowNum += 1

currentSchool = ''
SchoolInfo = {'President': [], 'Provost': [], 'Diversity': []}
#Iterate through all the elements in file
for row in range(1, 1026):
	institution = ws['A' + str(row)].value
	name = ws['B' + str(row)].value
	title = ws['C' + str(row)].value
	picture = ws['H' + str(row)].value
	gender = ''
	color = ws['B' + str(row)].fill.fgColor.rgb
	if color == 'FFFF00CC' or color == 'FFFF3399' or color == 'FFFFFF00' or color == 'FFFF0066':
		gender = 'Female'
	
	#when new school occurs
	if re.sub('[ \t\n\r-,&+]', '', currentSchool).lower() != re.sub('[ \t\n\r-,&+]', '', institution).lower():
		worksheet.write(rowNum, 1, currentSchool)
		while len(SchoolInfo['President']) != 0 or len(SchoolInfo['Provost']) != 0 or len(SchoolInfo['Diversity']) != 0:
			if len(SchoolInfo['President']) != 0:
				pres = SchoolInfo['President'].pop(0)
				worksheet.write(rowNum, 4, pres.name)
				worksheet.write(rowNum, 5, pres.gender)
				worksheet.write(rowNum, 7, pres.picture)
			if len(SchoolInfo['Provost']) != 0:
				prov = SchoolInfo['Provost'].pop(0)
				worksheet.write(rowNum, 8, prov.name)
				worksheet.write(rowNum, 9, prov.gender)
				worksheet.write(rowNum, 11, prov.picture)
			if len(SchoolInfo['Diversity']) != 0:
				div = SchoolInfo['Diversity'].pop(0)
				worksheet.write(rowNum, 12, div.name)
				worksheet.write(rowNum, 13, div.title)
				worksheet.write(rowNum, 14, div.gender)
				worksheet.write(rowNum, 16, div.picture)
			rowNum += 1
		#Update currentSchool to this one
		currentSchool = institution
		
	#Add this staff to currentSchool
	staff = Staff()
	staff.name = name
	staff.gender = gender
	staff.title = title
	staff.picture = picture
	
	if 'president' in title.lower() and 'vice' not in title.lower() and 'assistant' not in title.lower():
		SchoolInfo['President'].append(staff)
	if 'provost' in title.lower():
		SchoolInfo['Provost'].append(staff)
	if 'diversity' in title.lower():
		SchoolInfo['Diversity'].append(staff)
		
workbook.close()