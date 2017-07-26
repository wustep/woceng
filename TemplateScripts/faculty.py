import xlsxwriter, re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class Staff:
	name = ""
	gender = ""
	picture = ""
	dep = ""
	email = ""
	title = ""

#open xlsx file
print('Loading file...')
wb = load_workbook(filename='SimplifiedCFID.xlsx')
ws = wb['Sheet1']
print('Done loading file.')
#create new spreadsheet
workbook  = xlsxwriter.Workbook('Faculty.xlsx')
worksheet = workbook.add_worksheet()
print('Start writing to Faculty.xlsx')
rowNum = 0
#initial title setup
worksheet.write(rowNum, 0, 'IPEDS code')
worksheet.write(rowNum, 1, 'Institution')
worksheet.write(rowNum, 2, 'Department')
worksheet.write(rowNum, 3, 'Faculty Name')
worksheet.write(rowNum, 4, 'Faculty title (Assistant, Associate, or Full professor)')
worksheet.write(rowNum, 5, "Faculty Gender")
worksheet.write(rowNum, 6, "Faculty Ethnicity")
worksheet.write(rowNum, 7, "Faculty picture link (when available)")
worksheet.write(rowNum, 8, "Faculty Email")

rowNum += 1

currentSchool = ''
SchoolInfo = {}
#Iterate through all the elements in file
for row in range(1, 42569):
	institution = str(ws['A' + str(row)].value)
	name = ws['B' + str(row)].value
	title = ws['C' + str(row)].value
	dep = ws['D' + str(row)].value	
	email = ws['E' + str(row)].value
	gender = ws['K' + str(row)].value
	picture = ws['H' + str(row)].value

	
	#when new school occurs
	if re.sub('[ \t\n\r-,&+]', '', currentSchool).lower() != re.sub('[ \t\n\r-,&+]', '', institution).lower():
		worksheet.write(rowNum, 1, currentSchool)
		#Key is department name
		for key, value in SchoolInfo.items():
			worksheet.write(rowNum, 2, key)
			for elem in value:
				worksheet.write(rowNum, 3, elem.name)
				worksheet.write(rowNum, 4, elem.title)
				worksheet.write(rowNum, 5, elem.gender)
				worksheet.write(rowNum, 7, elem.picture)
				worksheet.write(rowNum, 8, elem.email)
				rowNum += 1
		SchoolInfo.clear()
		#Update currentSchool to this one
		currentSchool = institution
		
	#Add this staff to currentSchool
	staff = Staff()
	staff.name = name
	staff.gender = gender
	staff.title = title
	staff.picture = picture
	staff.dep = dep
	staff.email = email
	
	if dep not in SchoolInfo:
		SchoolInfo[dep] = []
	SchoolInfo[dep].append(staff)
	

		
workbook.close()