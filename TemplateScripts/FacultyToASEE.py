import xlsxwriter, re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from IPEDSSEARCH import IPEDSSEARCH

class Staff:
	name = ""
	gender = ""
	picture = ""
	dep = ""
	email = ""
	title = ""

#open xlsx file
print('Loading file...')
wb = load_workbook(filename='SimplifiedCFID.xlsx')
ws = wb['Sheet1']
print('Done loading file.')
#create new spreadsheet
workbook  = xlsxwriter.Workbook('Faculty.xlsx')
worksheet = workbook.add_worksheet()
print('Start writing to Faculty.xlsx')
rowNum = 0
#initial title setup
worksheet.write(rowNum, 0, 'IPEDS code')
worksheet.write(rowNum, 1, 'Institution')
worksheet.write(rowNum, 2, 'Department')
worksheet.write(rowNum, 3, 'Faculty Name')
worksheet.write(rowNum, 4, 'Faculty title (Assistant, Associate, or Full professor)')
worksheet.write(rowNum, 5, "Faculty Gender")
worksheet.write(rowNum, 6, "Faculty Ethnicity")
worksheet.write(rowNum, 7, "Faculty picture link (when available)")
worksheet.write(rowNum, 8, "Faculty Email")

rowNum += 1

currentSchool = ''
SchoolInfo = {}
newDB = IPEDSSEARCH()
newDB.init()

#Iterate through all the elements in file
for row in range(2, 2500):
	institution = str(ws['A' + str(row)].value)
	name = ws['B' + str(row)].value
	title = ws['C' + str(row)].value
	dep = ws['D' + str(row)].value	
	email = ws['E' + str(row)].value
	gender = ws['K' + str(row)].value
	picture = ws['H' + str(row)].value

	
	#when new school occurs
	if re.sub('[ \t\n\r-,&+]', '', currentSchool).lower() != re.sub('[ \t\n\r-,&+]', '', institution).lower():
		worksheet.write(rowNum, 1, currentSchool)
		IPEDS = newDB.IPEDSCode(re.sub('[ \t\n\r-,&+]', '', currentSchool).lower())
		worksheet.write(rowNum, 0, IPEDS)
		#Key is department name
		for key, value in SchoolInfo.items():
			worksheet.write(rowNum, 2, key)
			for elem in value:
				if not elem.name.strip():
					worksheet.write(rowNum, 3, "None")
				else:
					worksheet.write(rowNum, 3, elem.name)
				
				if not elem.title.strip():
					worksheet.write(rowNum, 4, "None")
				else:
					worksheet.write(rowNum, 4, elem.title)
				
				if not elem.gender.strip():
					worksheet.write(rowNum, 5, "None")
				else:
					worksheet.write(rowNum, 5, elem.gender)
				
				if not elem.picture.strip():
					worksheet.write(rowNum, 7, "None")
				else:
					worksheet.write(rowNum, 7, elem.picture)
				
				if not elem.email.strip():
					worksheet.write(rowNum, 8, "None")
				else:
					worksheet.write(rowNum, 8, elem.email)
				
				rowNum += 1
		SchoolInfo.clear()
		#Update currentSchool to this one
		currentSchool = institution
		
	#Add this staff to currentSchool
	staff = Staff()
	staff.name = str(name)
	staff.gender = str(gender)
	staff.title = str(title)
	staff.picture = str(picture)
	staff.dep = str(dep)
	staff.email = str(email)
	
	if dep not in SchoolInfo:
		SchoolInfo[dep] = []
	SchoolInfo[dep].append(staff)
	

		
workbook.close()